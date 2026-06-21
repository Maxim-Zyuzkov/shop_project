from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.contrib.auth.models import User, Group

import openpyxl
from io import BytesIO
from datetime import datetime

from .models import Product, Category, Manufacturer, Cart, CartItem, Order, OrderItem, Profile
from .serializers import ProductSerializer, CategorySerializer, ManufacturerSerializer, CartSerializer, CartItemSerializer, ProfileSerializer, OrderSerializer

# Импорты DRF
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser, BasePermission


# ==========================================
# КАСТОМНОЕ РАЗРЕШЕНИЕ
# ==========================================
class IsAdminOrReadOnly(BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        if request.user.is_authenticated:
            return request.user.groups.filter(name='Администраторы').exists()
        return False


# ==========================================
# ВЕБ-СТРАНИЦЫ
# ==========================================
def index(request):
    products = Product.objects.all().order_by('-id')[:6]
    categories = Category.objects.all()
    return render(request, 'shop/index.html', {'products': products, 'categories': categories})

def about_shop(request):
    return render(request, 'shop/about_shop.html')

def about_author(request):
    return render(request, 'shop/about_author.html')

def product_list(request):
    products = Product.objects.all()
    category_id = request.GET.get('category')
    if category_id: products = products.filter(category_id=category_id)
    manufacturer_id = request.GET.get('manufacturer')
    if manufacturer_id: products = products.filter(manufacturer_id=manufacturer_id)
    search_query = request.GET.get('q')
    if search_query: products = products.filter(Q(name__icontains=search_query) | Q(description__icontains=search_query))
    
    from django.core.paginator import Paginator
    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'shop/catalog.html', {
        'page_obj': page_obj,
        'categories': Category.objects.all(),
        'manufacturers': Manufacturer.objects.all(),
    })

def product_detail(request, pk):
    product = get_object_or_404(Product, id=pk)
    return render(request, 'shop/product_detail.html', {'product': product})

@login_required
def cart_view(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    return render(request, 'shop/cart.html', {'cart': cart})

@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_item, created = CartItem.objects.get_or_create(cart=cart, product=product)
    if not created:
        cart_item.quantity += 1
        cart_item.save()
        messages.success(request, f"Количество '{product.name}' увеличено!")
    else:
        messages.success(request, f"Товар '{product.name}' добавлен!")
    return redirect('cart_view')

@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    if request.method == 'POST':
        try:
            new_quantity = int(request.POST.get('quantity', 1))
            if new_quantity <= 0:
                cart_item.delete()
            else:
                cart_item.quantity = new_quantity
                cart_item.save()
        except ValueError: pass
    return redirect('cart_view')

@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    cart_item.delete()
    messages.success(request, "Товар удалён.")
    return redirect('cart_view')

@login_required
def checkout(request):
    cart = Cart.objects.filter(user=request.user).first()
    if not cart or not cart.items.exists():
        return redirect('cart_view')
    if request.method == 'POST':
        order = Order.objects.create(
            user=request.user,
            first_name=request.POST['first_name'],
            last_name=request.POST['last_name'],
            email=request.POST['email'],
            address=request.POST['address'],
            city=request.POST['city'],
            postal_code=request.POST['postal_code']
        )
        for cart_item in cart.items.all():
            OrderItem.objects.create(
                order=order, product=cart_item.product,
                price=cart_item.product.price, quantity=cart_item.quantity
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Чек по заказу'; ws['A2'] = f'Номер заказа: {order.id}'
        ws['A3'] = f'Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
        ws['A4'] = f'Покупатель: {order.first_name} {order.last_name}'
        ws['A8'] = '№'; ws['B8'] = 'Товар'; ws['C8'] = 'Цена'; ws['D8'] = 'Кол-во'; ws['E8'] = 'Сумма'
        row = 9
        for idx, item in enumerate(order.items.all(), start=1):
            ws[f'A{row}'] = idx; ws[f'B{row}'] = item.product.name
            ws[f'C{row}'] = float(item.price); ws[f'D{row}'] = item.quantity
            ws[f'E{row}'] = float(item.cost()); row += 1
        ws[f'D{row}'] = 'ИТОГО:'; ws[f'E{row}'] = float(order.total_cost())
        cart.delete()
        output = BytesIO(); wb.save(output); output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="check_{order.id}.xlsx"'
        messages.success(request, f"Заказ #{order.id} оформлен!")
        return response
    return render(request, 'shop/checkout.html', {'cart': cart})


# ==========================================
# РЕГИСТРАЦИЯ НОВОГО ПОЛЬЗОВАТЕЛЯ
# ==========================================
def register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email')
        full_name = request.POST.get('full_name')
        phone = request.POST.get('phone', '')
        city = request.POST.get('city', '')
        
        # Проверка, есть ли уже такой пользователь
        if User.objects.filter(username=username).exists():
            messages.error(request, "Пользователь с таким именем уже существует!")
            return redirect('register')
        
        # Создаём пользователя
        user = User.objects.create_user(username=username, password=password, email=email)
        
        # Создаём профиль и привязываем к пользователю
        profile = Profile.objects.create(
            user=user,
            full_name=full_name,
            phone=phone,
            city=city
        )
        
        # Автоматически добавляем в группу "Покупатели"
        buyer_group, created = Group.objects.get_or_create(name='Покупатели')
        user.groups.add(buyer_group)
        
        # Автоматический вход после регистрации
        login(request, user)
        messages.success(request, f"Добро пожаловать, {full_name}! Регистрация прошла успешно.")
        
        return redirect('index')
    
    return render(request, 'shop/register.html')


# ==========================================
# ЛИЧНЫЙ КАБИНЕТ
# ==========================================
@login_required
def profile_page(request):
    return render(request, 'shop/profile.html')


# ==========================================
# API
# ==========================================
@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def me_api(request):
    profile = request.user.profile
    if request.method == 'GET':
        serializer = ProfileSerializer(profile)
        return Response(serializer.data)
    elif request.method == 'PATCH':
        serializer = ProfileSerializer(profile, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_orders_api(request):
    orders = Order.objects.filter(user=request.user)
    serializer = OrderSerializer(orders, many=True)
    return Response(serializer.data)


# ==========================================
# VIEWSETS
# ==========================================
class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]

class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [IsAdminOrReadOnly]

class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    permission_classes = [IsAuthenticated]

class CartItemViewSet(viewsets.ModelViewSet):
    queryset = CartItem.objects.all()
    serializer_class = CartItemSerializer
    permission_classes = [IsAuthenticated]